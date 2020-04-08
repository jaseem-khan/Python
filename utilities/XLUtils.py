from openpyxl import Workbook
import openpyxl

class dataManipulation:

    def __init__(self, fileName, sheetName):
        self.fileName = fileName
        self.sheetName = sheetName
        self.wb = openpyxl.load_workbook(fileName)
        self.ws = self.wb["{}".format(sheetName)]

    def getRowCount(self):
        return self.ws.max_row

    def getColCount(self):
        return self.ws.max_column

    def readData(self, rowNo, columnNo):
        return self.ws.cell(rowNo, columnNo).value

    def writeData(self, rowNo, columnNo, data):
        self.ws.cell(rowNo, columnNo).value = data
        self.wb.save(self.fileName)

    def rowDataReader(self, min_row):

        data = []
        for row in self.ws.iter_rows(min_row=min_row, max_row=self.ws.max_row, max_col=self.ws.max_column):
            rowdata = []
            for cell in row:
                rowdata.append(cell.value)
            data.append(rowdata)

        return data







